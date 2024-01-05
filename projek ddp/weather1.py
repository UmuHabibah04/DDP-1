from tkinter import *
from datetime import date
from tkinter import messagebox
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pandas as pd

# Update the path accordingly
# Create a label with the image

background = "#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Form Registration System")
root.geometry("1250x700")
root.config(bg=background)


#Exit window

def Exit():
    root.destroy()

#reset 

def reset_fields():
    # Set all StringVar and IntVar variables to empty or zero
    Registration.set("")
    Name.set("")
    NIM.set(0)
    DOB.set("")
    radio.set(0)  # Assuming default value is 0 for no selection
    Class.set("Select Class")
    Religion.set("")

#gender

def save_data():
    registration_no = Registration.get()
    name = Name.get()
    nim = NIM.get()
    dob = DOB.get()
    gender = "Male" if radio.get() == 1 else "Female"
    class_val = Class.get()
    religion = Religion.get()
    

    # Append the data to the Excel file
    wb = openpyxl.load_workbook('Student_data.xlsx')
    sheet = wb.active

    next_row = sheet.max_row + 1

    sheet.cell(row=next_row, column=1).value = registration_no
    sheet.cell(row=next_row, column=2).value = name
    sheet.cell(row=next_row, column=3).value = nim
    sheet.cell(row=next_row, column=4).value = class_val
    sheet.cell(row=next_row, column=5).value = gender
    sheet.cell(row=next_row, column=6).value = dob
    sheet.cell(row=next_row, column=7).value = Date.get()  # Assuming Date is the registration date field
    sheet.cell(row=next_row, column=8).value = religion
    
    # Add other fields as necessary

    wb.save('Student_data.xlsx')
    messagebox.showinfo("Success", "Data saved successfully!")
#search box to update

Label1 = Label (root,  text="Form Register Student - System", font=("Halvetica", 20), fg=framebg,bg=background).place(x=400, y=30)
Search=StringVar()

Entry(root, textvariable=Search, width=15, bd=1, font="arial 20").place(x=815, y=135) 

Srch=Button(root, text="Search", compound=LEFT, width=12,bg='#68ddfa', font="arial 13 bold") 
Srch.place(x=1060, y=135)



#Registration and Date

Label(root, text="Registration No:", font="arial 13", fg=framebg,bg=background).place(x=30, y=150) 
Label(root, text="Date:", font="arial 13",fg=framebg,bg=background).place(x=500, y=150)

Registration=StringVar() 
Date = StringVar()

reg_entry= Entry (root, textvariable=Registration, width=15, font="arial 10") 
reg_entry.place(x=160, y=150)

#registration_no()

reg_entry= Entry (root, textvariable=Registration, width=15, font="arial 10") 
reg_entry.place(x=160,y=150)

#registration_no()

today= date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry= Entry (root, textvariable=Date, width=15, font="arial 10") 
date_entry.place(x=550,y=150)
Date.set(d1)

#Student
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=1200,bg=framebg,fg=framefg,height=250,relief=GROOVE) 
obj.place (x=30,y=250)

Label(obj, text="Full Name:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=50)
Label(obj, text="NIM:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100 )
Label(obj, text="Date of Birth:", font="arial 13", bg=framebg,fg=framefg).place(x=500,y=50) 
Label(obj, text="Gender:", font="arial 13", bg=framebg,fg=framefg).place(x=30, y=150)
Label(obj, text="Rombel:", font="arial 13", bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj, text="Religion:", font="arial 13", bg=framebg,fg=framefg).place( x=500, y=150)

Name=StringVar()
name_entry= Entry(obj, textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160, y=50)

NIM=IntVar()
nim_entry= Entry(obj, textvariable=NIM, width=20, font="arial 10")
nim_entry.place(x=160,y=100)

DOB = StringVar()
dob_entry= Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=630,y=50)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg,fg=framefg) 
R1.place(x=150,y=150)
R2= Radiobutton(obj,text="Female", variable=radio, value=2,bg=framebg,fg=framefg) 
R2.place(x=200,y=150) 

Religion = StringVar()
religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place( x=630,y=150)

Class = Combobox(obj,values=['TI01', 'TI02', 'TI03', 'TI04', 'TI05', 'TI06', 'TI07', 'TI08', 'TI09'], font="Roboto 10", width=17,state="r")
Class.place( x=630,y=100)
Class.set("Select Rombel")

#button

saveButton=Button(root, text="Submit", width=13, height=2, fg=framebg, font="arial 12 bold", bg="green", command=save_data) 
saveButton.place(x=250, y=520)

Button(root, text="Reset", width=13, height=2, fg=framebg, font="arial 12 bold", bg="red", command=reset_fields).place (x=550,y=520)
Button(root, text="Exit", width=13, height=2, fg=framebg, font="arial 12 bold", bg="grey", command=Exit).place(x=850,y=520)

root.mainloop()